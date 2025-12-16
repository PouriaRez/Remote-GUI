#!/usr/bin/env python3
"""
Power Monitoring Report Generator
Generates Excel and PDF reports from AnyLog power meter data.

Usage:
    python generate_power_report.py

Requirements:
    pip install pandas openpyxl pillow requests reportlab

For PDF conversion, reportlab is used for direct PDF generation.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import subprocess
import requests
import os

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as RLImage, Paragraph, Spacer
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors

# =============================================================================
# CONFIGURATION - Modify these values as needed
# =============================================================================
CONFIG = {
    'conn': '23.239.12.151:32349',
    'dbms': 'cos',
    'start_time': '2025-12-15 00:00:00',
    'end_time': '2025-12-16 00:00:00',
    'monitor_id': 'BG10',
    'increment_unit': 'hour',
    'increment_value': 1,
    'time_column': 'timestamp',
    'logo_path': "https://tse1.mm.bing.net/th/id/OIP.7bJT74xSx81aYJgz-rl-TwAAAA?cb=ucfimg2&ucfimg=1&rs=1&pid=ImgDetMain&o=7&rm=3",
    'engine_name': None,
    'engine_kw': '3000kW',
    'output_dir': 'outputs',
    'output_filename': 'power_monitoring_report2'
}
CONFIG["engine_name"] = f"Engine #{int(''.join(filter(str.isdigit, CONFIG['monitor_id'])))}"


# =============================================================================
# DATA FETCHING FUNCTIONS
# =============================================================================
def _get_data(conn: str, command: str, destination: str = None):
    """Execute AnyLog command and return response."""
    headers = {
        "command": command,
        "User-Agent": "AnyLog/1.23"
    }
    if destination:
        headers["destination"] = destination

    try:
        response = requests.get(url=f"http://{conn}", headers=headers)
        response.raise_for_status()
    except Exception as error:
        raise Exception(f"Failed to get data: {error}")
    return response


def _update_timestamp(end_time: str, increment_unit: str, increment_value: int) -> str:
    """Extend end timestamp by one increment to ensure complete coverage."""
    end_dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")

    if increment_unit == 'hour':
        end_dt += timedelta(hours=increment_value)
    elif increment_unit == 'day':
        end_dt += timedelta(days=increment_value)
    elif increment_unit == 'week':
        end_dt += timedelta(weeks=increment_value)
    elif increment_unit == 'minute':
        end_dt += timedelta(minutes=increment_value)
    else:
        end_dt += timedelta(hours=increment_value)

    return end_dt.strftime("%Y-%m-%d %H:%M:%S")


def image_download(image_url: str) -> str:
    """Download image from URL or return local path if file exists."""
    image_url_path = os.path.expandvars(os.path.expanduser(image_url))
    if os.path.isfile(image_url_path):
        return image_url_path

    try:
        content = requests.get(image_url).content
        local_path = 'downloaded_logo.png'
        with open(local_path, 'wb') as f:
            f.write(content)
        return local_path
    except Exception as error:
        print(f"Warning: Could not download image: {error}")
        return image_url_path


def check_data(conn: str, dbms: str) -> bool:
    """Validate database and table exist."""
    command = f"get data nodes where dbms={dbms} and table=pp_pm and format=json"
    try:
        response = _get_data(conn=conn, command=command)
        return bool(response.json())
    except Exception as e:
        print(f"Warning: Could not validate data source: {e}")
        return True  # Return True to allow data to be passed directly


def get_monitor_ids(conn: str, dbms: str) -> list:
    """Get list of available monitor IDs."""
    output = []
    query = "SELECT distinct(monitor_id) AS monitor_id FROM pp_pm"
    command = f"sql {dbms} format=json and stat=false {query}"

    try:
        response = _get_data(conn=conn, command=command, destination="network")
        for monitor_id in response.json()['Query']:
            output.append(monitor_id['monitor_id'])
    except Exception as error:
        print(f"Warning: Could not get monitor IDs: {error}")

    return output


def select_power_plant(conn: str, dbms: str, increment_unit: str, increment_value: int,
                       time_column: str, start_time: str, end_time: str, monitor_id: str) -> list:
    """Query power meter data from AnyLog."""
    query = f"""
    SELECT 
        increments({increment_unit}, {increment_value}, {time_column}), 
        monitor_id, 
        MIN(timestamp)::ljust(19) AS min_ts,
        AVG(realpower) AS kw, 
        AVG(powerfactor) AS pf, 
        AVG(a_current) AS amp_1, 
        AVG(b_current) AS amp_2, 
        AVG(c_current) AS amp_3,
        AVG(a_n_voltage) AS v_a,
        AVG(b_n_voltage) AS v_b,
        AVG(c_n_voltage) AS v_c
    FROM 
        pp_pm 
    WHERE
        {time_column} >= '{start_time}' AND 
        {time_column} < '{end_time}' AND 
        monitor_id='{monitor_id}' 
    GROUP BY 
        monitor_id 
    ORDER BY 
        min_ts
    """

    command = f"sql {dbms} format=json and stat=false {query.replace(chr(10), ' ').strip()}"

    try:
        response = _get_data(conn=conn, command=command, destination="network")
        return response.json()['Query']
    except Exception as error:
        raise Exception(f"Failed to query power plant data: {error}")


# =============================================================================
# PDF REPORT GENERATION (using reportlab)
# =============================================================================
def generate_pdf_report(df: pd.DataFrame, config: dict) -> str:
    """Generate the Power Monitoring PDF report using reportlab."""
    pdf_path = os.path.join(config['output_dir'], f"{config['output_filename']}.pdf")
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(letter),
        leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20
    )

    elements = []
    styles = getSampleStyleSheet()

    # Create centered style for subtitle
    centered_style = ParagraphStyle(
        'Centered',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=11,
        spaceAfter=8
    )

    # Create custom title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=6
    )

    # Timestamp and Logo on same line
    now = datetime.now()
    timestamp = f"{now.month}/{now.day}/{now.year} {now.strftime('%H:%M:%S')}"

    if os.path.exists(config["logo_path"]):
        # Create a table to align timestamp (left) and logo (right)
        logo = RLImage(config["logo_path"])
        logo.drawHeight = 40
        logo.drawWidth = 40

        header_table = Table([[Paragraph(timestamp, styles["Normal"]), logo]], colWidths=[700, 50])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header_table)
    else:
        elements.append(Paragraph(timestamp, styles["Normal"]))

    elements.append(Spacer(1, 8))

    # Title (centered)
    elements.append(Paragraph(
        "<b>CITY OF SABETHA MUNICIPAL POWER PLANT</b>",
        title_style
    ))

    # Subtitle (centered)
    elements.append(Paragraph(
        "<b>EVERGY INTERCONNECTION</b>",
        centered_style
    ))
    elements.append(Spacer(1, 6))

    # Engine name and kW rating
    engine_info = f"<b>{config.get('engine_name', '')} - {config.get('engine_kw', '')}</b>"
    elements.append(Paragraph(engine_info, centered_style))
    elements.append(Spacer(1, 6))

    # Data Table - AMPS and VOLTAGE in row 0, other labels in row 1
    table_data = []

    # Row 0: Span columns with labels
    table_data.append([
        "", "", "", "AMPS", "", "",
        "VOLTAGE", "", ""
    ])

    # Row 1: All column sub-headers
    table_data.append([
        "DATE/TIME", "KW", "PF", "1", "2", "3",
        "A", "B", "C"
    ])

    # Data rows
    for _, row in df.iterrows():
        ts = pd.to_datetime(row["min_ts"]).floor("h")
        # Format date without leading zeros
        datetime_str = f"{ts.month}/{ts.day}/{ts.year} {ts.strftime('%I:%M:%S %p')}"

        table_data.append([
            datetime_str,
            int(round(row["kw"])) if pd.notna(row["kw"]) else 0,
            int(round(row["pf"] / 100)) if pd.notna(row["pf"]) else 0,
            int(round(row["amp_1"])) if pd.notna(row["amp_1"]) else 0,
            int(round(row["amp_2"])) if pd.notna(row["amp_2"]) else 0,
            int(round(row["amp_3"])) if pd.notna(row["amp_3"]) else 0,
            int(round(row["v_a"])) if pd.notna(row["v_a"]) else 0,
            int(round(row["v_b"])) if pd.notna(row["v_b"]) else 0,
            int(round(row["v_c"])) if pd.notna(row["v_c"]) else 0
        ])

    # Create table with comfortable sizing
    table = Table(table_data, repeatRows=2)

    table_style = TableStyle([
        # Row 0: Labels spanning - gray background
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("SPAN", (0, 0), (0, 0)),  # DATE/TIME
        ("SPAN", (1, 0), (1, 0)),  # KW
        ("SPAN", (2, 0), (2, 0)),  # PF
        ("SPAN", (3, 0), (5, 0)),  # AMPS spans 3 columns
        ("SPAN", (6, 0), (8, 0)),  # VOLTAGE spans 3 columns
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),

        # Row 1: Sub-headers - gray background
        ("BACKGROUND", (0, 1), (-1, 1), colors.lightgrey),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.black),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 8),
        ("ALIGN", (0, 1), (-1, 1), "CENTER"),

        # Data rows styling
        ("ALIGN", (1, 2), (-1, -1), "CENTER"),
        ("ALIGN", (0, 2), (0, -1), "LEFT"),
        ("FONTSIZE", (0, 2), (-1, -1), 8),

        # Comfortable padding for all cells
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),

        # Grid and borders
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),

        # THICKER LINE between row 1 and row 2 (between headers and data)
        ("LINEBELOW", (0, 1), (-1, 1), 2, colors.black),

        # THICKER LINE between column 0 and column 1 (DATE/TIME and KW)
        ("LINEAFTER", (0, 0), (0, -1), 2, colors.black),
    ])
    table.setStyle(table_style)
    table.hAlign = 'LEFT'  # Left-justify the table

    elements.append(table)

    # Footer section - 3 columns, 7 rows (2 blank lines above)
    elements.append(Spacer(1, 24))  # 2 blank lines

    footer_data = [
        ["DAILY READING", "PREVIOUS", "PRESENT", "UNITS"],
        ["KHW IN", "", "", ""],
        ["TIME METER", "", "", ""],
        ["FUEL OIL METER", "", "", ""],
        ["GAS EAST METER", "", "", ""],
        ["S. PLANT KWHS", "", "", ""],
        ["LUBE OIL", "", "", ""]
    ]

    footer_table = Table(footer_data, colWidths=[120, 90, 90, 90])
    footer_table_style = TableStyle([
        # Header row (row 0)
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (-1, 0), "CENTER"),

        # First column (labels) - bold, left-aligned
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (0, -1), colors.lightgrey),

        # Other columns - centered
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),

        # All cells styling
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 1, colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    footer_table.setStyle(footer_table_style)
    footer_table.hAlign = 'LEFT'  # Left-justify the footer table

    elements.append(footer_table)

    doc.build(elements)

    print(f"PDF file created: {pdf_path}")
    return pdf_path


# =============================================================================
# EXCEL REPORT GENERATION
# =============================================================================
def generate_excel_report(df: pd.DataFrame, config: dict) -> str:
    """Generate the Power Monitoring Excel report."""

    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    # Page setup for single-page printing
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

    # Define styles
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    bold_font = Font(bold=True)

    # Row 1: Timestamp (left) and Logo (right)
    ws['A1'] = datetime.now().strftime('%m/%d/%Y %H:%M:%S')
    ws['A1'].alignment = left_align

    # Add logo if exists
    logo_path = config.get('logo_path', '')
    if logo_path and os.path.exists(logo_path):
        try:
            logo = Image(logo_path)
            logo.width = 60
            logo.height = 60
            ws.add_image(logo, 'I1')
        except Exception as e:
            print(f"Warning: Could not add logo: {e}")

    # Row 2: Title
    ws.merge_cells('A2:I2')
    ws['A2'] = 'CITY OF SABETHA MUNICIPAL POWER PLANT'
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = center_align

    # Row 3: Subtitle
    ws.merge_cells('A3:I3')
    ws['A3'] = 'EVERGY INTERCONNECTION'
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = center_align

    # Row 4: Empty (spacing)

    # Row 5: Engine name (left) and kW rating (right)
    ws['B5'] = config.get('engine_name', '')
    ws['B5'].font = Font(bold=True, size=14)
    ws['B5'].alignment = left_align

    ws['I5'] = config.get('engine_kw', '')
    ws['I5'].font = Font(bold=True, size=14)
    ws['I5'].alignment = right_align

    # Row 6: Category headers (gray background)
    for col in ['A', 'B', 'C']:
        ws[f'{col}6'].fill = gray_fill
        ws[f'{col}6'].border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # AMPS header (D-F)
    ws['D6'] = 'AMPS'
    ws['D6'].font = bold_font
    ws['D6'].alignment = center_align
    for col in ['D', 'E', 'F']:
        ws[f'{col}6'].fill = gray_fill
        ws[f'{col}6'].border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    ws.merge_cells('D6:F6')

    # VOLTAGE header (G-I)
    ws['G6'] = 'VOLTAGE'
    ws['G6'].font = bold_font
    ws['G6'].alignment = center_align
    for col in ['G', 'H', 'I']:
        ws[f'{col}6'].fill = gray_fill
        ws[f'{col}6'].border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    ws.merge_cells('G6:I6')

    # Row 7: Column headers
    headers = ['DATE/TIME', 'KW', 'PF', '1', '2', '3', 'A', 'B', 'C']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = gray_fill
        left_style = medium_side if col_num == 1 else thin_side
        right_style = medium_side if col_num in [1, 9] else thin_side
        cell.border = Border(left=left_style, right=right_style, top=medium_side, bottom=medium_side)

    # Data rows (starting row 8)
    data_start_row = 8
    num_rows = len(df)

    for idx, row in df.iterrows():
        data_row = data_start_row + idx
        is_last_row = (idx == num_rows - 1)

        # Format timestamp - floor to hour, 12-hour format
        ts = pd.to_datetime(row['min_ts']).floor('h')
        datetime_str = f"{ts.month}/{ts.day}/{ts.year} {ts.strftime('%I:%M:%S %p')}"

        # Values as whole numbers
        kw_val = int(round(row['kw'])) if pd.notna(row['kw']) else 0
        pf_val = int(round(row['pf'] / 100)) if pd.notna(row['pf']) else 0
        amp_1 = int(round(row['amp_1'])) if pd.notna(row['amp_1']) else 0
        amp_2 = int(round(row['amp_2'])) if pd.notna(row['amp_2']) else 0
        amp_3 = int(round(row['amp_3'])) if pd.notna(row['amp_3']) else 0
        v_a = int(round(row['v_a'])) if pd.notna(row['v_a']) else 0
        v_b = int(round(row['v_b'])) if pd.notna(row['v_b']) else 0
        v_c = int(round(row['v_c'])) if pd.notna(row['v_c']) else 0

        values = [datetime_str, kw_val, pf_val, amp_1, amp_2, amp_3, v_a, v_b, v_c]

        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_num)
            cell.value = val
            left_style = medium_side if col_num == 1 else thin_side
            right_style = medium_side if col_num in [1, 9] else thin_side
            bottom_style = medium_side if is_last_row else thin_side
            cell.border = Border(left=left_style, right=right_style, top=thin_side, bottom=bottom_style)
            cell.alignment = center_align

    # Footer section
    footer_start = data_start_row + num_rows

    # First footer row: DAILY READING with headers
    footer_row = footer_start
    ws.cell(row=footer_row, column=1).value = 'DAILY READING'
    ws.cell(row=footer_row, column=1).font = bold_font
    ws.cell(row=footer_row, column=1).alignment = right_align
    ws.cell(row=footer_row, column=1).border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # PREVIOUS (B-C)
    ws.cell(row=footer_row, column=2).value = 'PREVIOUS'
    ws.cell(row=footer_row, column=2).font = bold_font
    ws.cell(row=footer_row, column=2).alignment = center_align
    for col in [2, 3]:
        ws.cell(row=footer_row, column=col).border = Border(left=thin_side, right=thin_side, top=thin_side,
                                                            bottom=thin_side)
    ws.merge_cells(f'B{footer_row}:C{footer_row}')

    # PRESENT (D-F)
    ws.cell(row=footer_row, column=4).value = 'PRESENT'
    ws.cell(row=footer_row, column=4).font = bold_font
    ws.cell(row=footer_row, column=4).alignment = center_align
    for col in [4, 5, 6]:
        ws.cell(row=footer_row, column=col).border = Border(left=thin_side, right=thin_side, top=thin_side,
                                                            bottom=thin_side)
    ws.merge_cells(f'D{footer_row}:F{footer_row}')

    # UNITS (G-I)
    ws.cell(row=footer_row, column=7).value = 'UNITS'
    ws.cell(row=footer_row, column=7).font = bold_font
    ws.cell(row=footer_row, column=7).alignment = center_align
    for col in [7, 8, 9]:
        ws.cell(row=footer_row, column=col).border = Border(left=thin_side, right=thin_side, top=thin_side,
                                                            bottom=thin_side)
    ws.merge_cells(f'G{footer_row}:I{footer_row}')

    # Footer labels
    footer_labels = ['KHW IN', 'TIME METER', 'FUEL OIL METER', 'GAS EAST METER', 'S. PLANT KWHS', 'LUBE OIL']
    for i, label in enumerate(footer_labels):
        row = footer_start + 1 + i
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = bold_font
        ws.cell(row=row, column=1).alignment = right_align
        ws.cell(row=row, column=1).border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for col in [2, 3]:
            ws.cell(row=row, column=col).border = Border(left=thin_side, right=thin_side, top=thin_side,
                                                         bottom=thin_side)
        ws.merge_cells(f'B{row}:C{row}')

        for col in [4, 5, 6]:
            ws.cell(row=row, column=col).border = Border(left=thin_side, right=thin_side, top=thin_side,
                                                         bottom=thin_side)
        ws.merge_cells(f'D{row}:F{row}')

        for col in [7, 8, 9]:
            ws.cell(row=row, column=col).border = Border(left=thin_side, right=thin_side, top=thin_side,
                                                         bottom=thin_side)
        ws.merge_cells(f'G{row}:I{row}')

    # Column widths
    widths = {'A': 24, 'B': 8, 'C': 6, 'D': 6, 'E': 6, 'F': 6, 'G': 8, 'H': 8, 'I': 8}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Save Excel file
    os.makedirs(config['output_dir'], exist_ok=True)
    excel_path = os.path.join(config['output_dir'], f"{config['output_filename']}.xlsx")
    wb.save(excel_path)
    print(f"Excel file saved: {excel_path}")

    return excel_path


def main(data: list = None):
    """Main function to generate the report.

    Args:
        data: Optional list of power meter data dictionaries. If provided, skips REST API query.
    """
    config = CONFIG.copy()

    print("Generating Power Monitoring Report...")
    print(f"Date Range: {config['start_time']} to {config['end_time']}")
    print(f"Monitor ID: {config['monitor_id']}")
    print(f"Engine: {config['engine_name']} ({config['engine_kw']})")
    print("-" * 50)

    if data is None:
        # Validate data source
        check_data(conn=config['conn'], dbms=config['dbms'])

        # Extend end time by one increment
        extended_end_time = _update_timestamp(
            config['end_time'],
            config['increment_unit'],
            config['increment_value']
        )

        # Query power plant data
        try:
            pp_data = select_power_plant(
                conn=config['conn'],
                dbms=config['dbms'],
                increment_unit=config['increment_unit'],
                increment_value=config['increment_value'],
                time_column=config['time_column'],
                start_time=config['start_time'],
                end_time=extended_end_time,
                monitor_id=config['monitor_id']
            )
        except Exception as e:
            print(f"Error querying data: {e}")
            print("Please provide data directly using main(data=[...])")
            return None
    else:
        pp_data = data

    if not pp_data:
        raise ValueError("No data returned from query")

    print(f"Retrieved {len(pp_data)} data points")

    # Create dataframe
    df = pd.DataFrame(pp_data)

    # Download logo if it's a URL
    if config['logo_path'].startswith('http'):
        config['logo_path'] = image_download(config['logo_path'])

    # Generate Excel report
    excel_path = generate_excel_report(df, config)

    # Generate PDF report using reportlab (instead of LibreOffice conversion)
    pdf_path = generate_pdf_report(df, config)

    print("-" * 50)
    print("Report generation complete!")

    return excel_path, pdf_path


if __name__ == '__main__':
    main()